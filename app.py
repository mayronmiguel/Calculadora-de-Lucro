import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl

PLANILHA = "planilha.xlsx"


# ------------------------------------------
# Função: Obter dados da planilha
# ------------------------------------------
def obter_dados_excel():
    wb = openpyxl.load_workbook(PLANILHA)
    sheet = wb.active
    dados = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        dados.append(row)
    return dados


# ------------------------------------------
# Função: Deletar produto da planilha
# ------------------------------------------
def deletar_produto(nome_produto):
    wb = openpyxl.load_workbook(PLANILHA)
    sheet = wb.active
    contador = 2

    for row in sheet.iter_rows(min_row=2, min_col=1, values_only=True):
        if str(row[0]) == nome_produto:
            sheet.delete_rows(contador)
            wb.save(PLANILHA)
            return True
        contador += 1
    return False


# ------------------------------------------
# Função: Atualizar tabela gráfica
# ------------------------------------------
def atualizar_tabela():
    for linha in tabela.get_children():
        tabela.delete(linha)

    dados = obter_dados_excel()

    for item in dados:
        tabela.insert("", tk.END, values=item)


# ------------------------------------------
# Função: Ação do botão excluir
# ------------------------------------------
def acao_excluir():
    item_selecionado = tabela.focus()

    if not item_selecionado:
        messagebox.showwarning("Aviso", "Selecione um produto para excluir.")
        return

    valores = tabela.item(item_selecionado, "values")
    nome_produto = valores[0]

    confirmar = messagebox.askyesno(
        "Confirmar",
        f"Deseja realmente excluir o produto '{nome_produto}'?"
    )

    if confirmar:
        if deletar_produto(nome_produto):
            messagebox.showinfo("Sucesso", "Produto excluído com sucesso!")
            atualizar_tabela()
        else:
            messagebox.showerror("Erro", "Não foi possível excluir o produto.")


# ------------------------------------------
# CRIAÇÃO DA JANELA PRINCIPAL
# ------------------------------------------
janela = tk.Tk()
janela.title("Calculadora de Lucro - Visualizador")
janela.geometry("900x400")
janela.configure(bg="#f0f0f0")


# ------------------------------------------
# Título
# ------------------------------------------
titulo = tk.Label(
    janela,
    text="📊 Produtos Registrados na Planilha",
    font=("Arial", 18, "bold"),
    bg="#f0f0f0"
)
titulo.pack(pady=10)


# ------------------------------------------
# Tabela Treeview
# ------------------------------------------
colunas = (
    "Produto", "Compra", "Venda", "Custos", "Frete",
    "Lucro", "Margem (%)"
)

tabela = ttk.Treeview(janela, columns=colunas, show="headings", height=15)

for col in colunas:
    tabela.heading(col, text=col)
    tabela.column(col, width=120)

tabela.pack(padx=10, pady=10, fill=tk.BOTH, expand=True)

# Adiciona uma barra de rolagem
scroll = ttk.Scrollbar(janela, orient="vertical", command=tabela.yview)
tabela.configure(yscroll=scroll.set)
scroll.pack(side=tk.RIGHT, fill=tk.Y)


# ------------------------------------------
# Botão deletar
# ------------------------------------------
btn_excluir = ttk.Button(
    janela,
    text="Excluir Produto Selecionado",
    command=acao_excluir
)
btn_excluir.pack(pady=10)


# ------------------------------------------
# Carrega os dados ao iniciar
# ------------------------------------------
atualizar_tabela()


# ------------------------------------------
# Iniciar a interface
# ------------------------------------------
janela.mainloop()
