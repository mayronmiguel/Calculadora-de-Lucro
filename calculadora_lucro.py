# importando openpyxl
import openpyxl

# definindo as variaveis 

preco_compra = float(input('Digite o preco de compra: '))
preco_venda = float(input('Digite o preco de venda: '))
custos_adicionais = float(input('Digite os custos adicionais (opcional): '))
custo_frete = float(input('Digite o custo medio de frete: '))

# calculando o lucro liquido

lucro = preco_venda - preco_compra - custos_adicionais - custo_frete

# imprimindo o resultado
print("O lucro liquido e de R${:.2f}".format(lucro))

# calculando a margem de lucro
margem_lucro = (lucro/preco_venda) * 100

# imprimindo o resultado
print("A margem de lucro e de {:.2f}".format(margem_lucro))

# salvando resultados em uma planilha do excel 

nome_do_produto = 'Produto 1'

resumo = [nome_do_produto, preco_compra, preco_venda, custos_adicionais, custo_frete, lucro, margem_lucro]

# carregando a planilha existente ou criando uma nova planilha se ela não existir
try: 
    wb =openpyxl.load_workbook('planilha.xlsx')
    sheet = wb.active
except:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Resultado da Calculadora de Lucro"

    # Adicionando cabecalho 
    sheet['A1'] = "Nome do Produto"
    sheet['B1'] = "Preco da Compra"
    sheet['C1'] = "Preco de Venda"
    sheet['D1'] = "Custos Adicionais"
    sheet['E1'] = "Custo Medio de Frete"
    sheet['F1'] = "Lucro Líquido"
    sheet['G1'] = "Margem de Lucro (%)"

# Adicionando valores a planilha 
row = sheet.max_row + 1
sheet['A{}'.format(row)] = nome_do_produto
sheet['B{}'.format(row)] = preco_compra
sheet['C{}'.format(row)] = preco_venda
sheet['D{}'.format(row)] = custos_adicionais
sheet['E{}'.format(row)] = custo_frete
sheet['F{}'.format(row)] = lucro
sheet['G{}'.format(row)] = margem_lucro

# salvando a planilha 
wb.save('planilha.xlsx')